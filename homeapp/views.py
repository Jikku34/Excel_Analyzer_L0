from django.shortcuts import render, redirect
from .models import ExcelFile, StudentAdmission, StudentFee, StudentRegistration, StudentFeepaid

from openpyxl import load_workbook


# function used to check the user is existed

def home(request):
    if 'user' in request.session:

        return redirect('show')

    else:

        return render(request, 'index.html')


# function for index page and save excel data into database
def index(request):

    if request.method == 'POST':

        if 'registration_file_upload' in request.FILES:
            request.session['user'] = "file_exit"
            data = ExcelFile()
            data.exel_file = request.FILES['registration_file_upload']
            A = str(request.FILES['registration_file_upload'].name)
            d = name_file(A)

            data.filename = d
            data.save()
            # +++++++++++++++++++++++++
            data1 = ExcelFile()
            data1.exel_file = request.FILES['admission_file_upload']
            B = str(request.FILES['admission_file_upload'].name)
            c = name_file(B)
            print(c)

            data1.filename = c
            data1.save()
            # +++++++++++++++++++++++++++
            data2 = ExcelFile()
            data2.exel_file = request.FILES['fee_file_upload']
            C = str(request.FILES['fee_file_upload'].name)
            e = name_file(C)
            data2.filename = e
            data2.save()

            # save fee excel file data into database

            wb = load_workbook('C:\\Users\\DELL\\PycharmProjects\\ExcelTaskL0\\media\\excelfile\\' + e,
                               data_only=True)
            sheet = wb.active

            j = 2
            while j < 150:
                if sheet['C' + str(j)].value is None:
                    x = 0

                else:
                    data1 = StudentFee()
                    data1.student_date_fee = sheet['A' + str(j)].value
                    data1.student_id = sheet['B' + str(j)].value
                    data1.student_name = sheet['C' + str(j)].value
                    data1.student_fee = sheet['D' + str(j)].value
                    data1.save()
                j = j + 1

            # save the registration excel data into database

            wc = load_workbook('C:\\Users\\DELL\\PycharmProjects\\ExcelTaskL0\\media\\excelfile\\' + d,
                               data_only=True)
            sheet = wc.active

            j = 3
            while j < 150:
                if sheet['D' + str(j)].value is None:
                    x = 0

                else:
                    data1 = StudentRegistration()
                    data1.student_reg_date = sheet['A' + str(j)].value
                    data1.student_reg_no = sheet['B' + str(j)].value
                    data1.branch = sheet['C' + str(j)].value
                    data1.student_name = sheet['D' + str(j)].value
                    data1.student_phone = sheet['E' + str(j)].value
                    data1.student_email = sheet['F' + str(j)].value
                    data1.student_address = sheet['G' + str(j)].value
                    data1.student_dist = sheet['H' + str(j)].value
                    data1.student_state = sheet['I' + str(j)].value
                    data1.student_company = sheet['J' + str(j)].value
                    data1.student_specialisation = sheet['K' + str(j)].value
                    data1.student_guardian_name = sheet['L' + str(j)].value
                    data1.student_guardian_phone = sheet['M' + str(j)].value
                    data1.student_enquiry_Source = sheet['N' + str(j)].value
                    data1.student_referral = sheet['O' + str(j)].value
                    data1.student_staff = sheet['P' + str(j)].value
                    data1.student_department = sheet['Q' + str(j)].value
                    data1.student_Course = sheet['R' + str(j)].value
                    data1.student_effective_fee = sheet['S' + str(j)].value
                    data1.student_reg_fee = sheet['T' + str(j)].value

                    data1.save()
                j = j + 1

            #  save the admission excel data into database

            wc = load_workbook('C:\\Users\\DELL\\PycharmProjects\\ExcelTaskL0\\media\\excelfile\\' + c,
                               data_only=True)
            sheet = wc.active

            j = 3
            while j < 150:
                if sheet['D' + str(j)].value is None:
                    x = 0

                else:
                    data1 = StudentAdmission()
                    data1.student_admission_date = sheet['A' + str(j)].value
                    data1.student_admission_no = sheet['B' + str(j)].value
                    data1.student_reg_date = sheet['C' + str(j)].value
                    data1.student_reg_no = sheet['D' + str(j)].value
                    data1.branch = sheet['E' + str(j)].value
                    data1.student_name = sheet['F' + str(j)].value
                    data1.student_phone = sheet['G' + str(j)].value
                    data1.student_email = sheet['H' + str(j)].value
                    data1.student_address = sheet['I' + str(j)].value
                    data1.student_dist = sheet['J' + str(j)].value
                    data1.student_state = sheet['K' + str(j)].value
                    data1.student_company = sheet['L' + str(j)].value
                    data1.student_specialisation = sheet['M' + str(j)].value
                    data1.student_guardian_name = sheet['N' + str(j)].value
                    data1.student_guardian_phone = sheet['O' + str(j)].value
                    data1.student_enquiry_Source = sheet['P' + str(j)].value
                    data1.student_referral = sheet['Q' + str(j)].value
                    data1.student_staff = sheet['R' + str(j)].value
                    data1.student_department = sheet['S' + str(j)].value
                    data1.student_Course = sheet['T' + str(j)].value
                    data1.student_effective_fee = sheet['U' + str(j)].value
                    data1.student_mode_training = sheet['V' + str(j)].value
                    data1.student_start_date = sheet['W' + str(j)].value
                    data1.save()
                j = j + 1
            amout_save()

            return render(request, 'show.html')
        elif 'registration_file_upload' not in request.FILES:
            if 'user' in request.session:
                return redirect('show')
            else:
                return render(request, 'index.html')

        else:
            render(request, 'index.html')
    else:
        return render(request, 'index.html')

    return render(request, 'index.html')


# function for handle, show , filter the Excel database data
def show(request):
    if 'user' in request.session:
        if request.method == 'POST':
            category = request.POST.get('Category')

            if category is None:
                d = 'start'
                return render(request, 'show.html', {'d': d})

            elif category == "admission":
                d = "admission"
                data = StudentAdmission.objects.all()
                return render(request, 'show.html', {'data': data, 'd': d})

            elif category == "fee":
                d = "fees"
                d2 = "startfee"
                data13 = StudentAdmission.objects.all()
                data14 = StudentFeepaid.objects.all()
                if request.method == 'POST':
                    fee_category = request.POST.get('fee_category')
                    search = request.POST.get('search')

                    # check the value of search box value

                    if search:
                        data20 = StudentFeepaid.objects.filter(student_id=search).values()

                        data21 = StudentAdmission.objects.filter(student_reg_no=search).values()

                        # check or compare the value is student registration number

                        if data20:
                            return render(request, 'show.html', {'data13': data21, 'data14': data20, 'd': d})

                        # check and compare the search box value to student name

                        else:
                            # compare the search value with like operation
                            result = StudentFeepaid.objects.filter(student_name__contains=search).values()
                            # if result has value
                            if result:
                                d1 = 'search'
                                data21 = StudentAdmission.objects.all()

                                return render(request, 'show.html',
                                              {'data13': data21, 'data14': result, 'd': d, 'd1': d1})

                    else:
                        # showing the fee history of students

                        if fee_category == "Fee_History":
                            d1 = "Fee_History"
                            data = StudentAdmission.objects.all()
                            data1 = StudentFee.objects.all()
                            return render(request, 'show.html', {'data': data, 'data1': data1, 'd': d, 'd1': d1})

                        # showing the fee full payment of students

                        elif fee_category == "Full_Payment_Students":
                            d1 = "Full_Payment_Students"
                            data9 = StudentAdmission.objects.all()
                            data10 = StudentFeepaid.objects.filter(student_fee_status="PAID").values()
                            return render(request, 'show.html', {'data9': data9, 'data10': data10, 'd': d, 'd1': d1})

                        # showing the fee pedding fee of students

                        elif fee_category == "Pending_Fee":
                            d1 = "Pending_Fee"
                            data7 = StudentAdmission.objects.all()
                            data8 = StudentFeepaid.objects.filter(student_fee_status="PENDING").values()
                            return render(request, 'show.html', {'data7': data7, 'data8': data8, 'd': d, 'd1': d1})

                # show the fee details of student default

                return render(request, 'show.html', {'data13': data13, 'data14': data14, 'd': d})

            # details of student registration

            else:
                d = "registration"
                data = StudentRegistration.objects.all()
                return render(request, 'show.html', {'data': data, 'd': d})

    else:
        return render(request, 'index.html')
    return render(request, 'show.html')


# fun for remove the white space add underscore
def name_file(A):
    i = 0
    f = 0

    new = list(A)
    while i < len(A):
        if new[i] == " ":
            new[i] = "_"
        else:
            f = 0
        i = i + 1

    d = ''.join(new)
    return d


# function for find the sum of total fee , total payments,effective fee, and stored into database

def amout_save():
    d = 0
    data = StudentAdmission.objects.all()
    data1 = StudentFee.objects.all()
    for x in data:
        for y in data1:
            if x.student_reg_no == y.student_id:
                if y.student_fee:
                    d = d + y.student_fee
                else:
                    f = 0
            else:
                f = 0
        c = d

        b = x.student_effective_fee - d
        d = 0
        data4 = StudentFeepaid()
        data4.student_id = x.student_reg_no
        data4.student_name = x.student_name
        data4.student_effective_fee = x.student_effective_fee
        data4.student_fee_paid = c
        data4.student_balance_fee = b
        data4.student_Course = x.student_Course
        if b == 0 or b < 0:
            data4.student_fee_status = "PAID"
        else:
            data4.student_fee_status = "PENDING"

        data4.save()

# function for student dashboard

def student_profile(request, id):
    dts = StudentAdmission.objects.filter(student_reg_no=id)
    dts1 = StudentRegistration.objects.filter(student_reg_no=id)
    if request.method == 'POST':
        student_category = request.POST.get('student_category')

        # condition to check the admission details of  current student

        if student_category == "admission_details":
            d = "admission_details"

            if dts:
                return render(request, 'student_profile.html', {'dt': dts, 'd': d})
            else:
                return render(request, 'student_profile.html', {'dt': dts1, 'd': d})

        # condition to check the fee details of  current student

        elif student_category == "fee_details":

            d = "fee_details"
            dts = StudentFee.objects.filter(student_id=id)
            dts1 = StudentFeepaid.objects.filter(student_id=id)
            return render(request, 'student_profile.html', {'dt': dts, 'dt1': dts1, 'd': d})

        # condition to check the profile details of  current student

        else:
            d = 'profile'
            if dts:
                return render(request, 'student_profile.html', {'dt': dts, 'd': d})

            else:
                return render(request, 'student_profile.html', {'dt': dts1, 'd': d})

    # default make profile of current student

    d = 'profile'
    if dts:
        return render(request, 'student_profile.html', {'dt': dts, 'd': d})

    else:
        return render(request, 'student_profile.html', {'dt': dts1, 'd': d})


# function to remove all data form database start new analyse

def remove(request):
    try:
        ExcelFile.objects.all().delete()
        StudentRegistration.objects.all().delete()
        StudentAdmission.objects.all().delete()
        StudentFee.objects.all().delete()
        StudentFeepaid.objects.all().delete()
        del request.session['user']

    except:
        return redirect('home-page')
    return redirect('home-page')
